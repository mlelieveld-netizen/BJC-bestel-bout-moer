#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract fasteners (bouten, moeren, schroeven) uit Excel en maak JSON
"""

import openpyxl
import json
import re

filename = "BJC bestel artikelen Oetelaar (1).xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb.active

# Keywords voor fasteners
fastener_keywords = [
    'bout', 'moer', 'schroef', 'schroeven', 'ring', 'ringetje', 'ringetjes',
    'kram', 'krammen', 'nagel', 'nagels', 'spijker', 'spijkers',
    'dop', 'doppen', 'plug', 'plugs', 'anker', 'ankers',
    'clips', 'clip', 'klem', 'klemmen', 'beugel', 'beugels',
    'fastener', 'fasteners', 'verbinding', 'verbindingsmateriaal',
    'm6', 'm8', 'm10', 'm12', 'm16', 'm20',  # Metrische maten
    'din', 'iso', 'unf', 'unc'  # Standaarden
]

def is_fastener(omschrijving):
    """Check of artikelomschrijving een fastener is"""
    if not omschrijving:
        return False
    omschrijving_lower = str(omschrijving).lower()
    return any(keyword in omschrijving_lower for keyword in fastener_keywords)

# Lees headers
headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(1, col).value)

print("=" * 70)
print("EXTRACT FASTENERS UIT EXCEL")
print("=" * 70)

fasteners = []
fastener_count = 0

# Lees alle rijen
for row in range(2, ws.max_row + 1):
    artikelnummer = ws.cell(row, 1).value
    omschrijving = ws.cell(row, 2).value
    brutoprijs = ws.cell(row, 3).value
    korting = ws.cell(row, 4).value
    nettoprijs = ws.cell(row, 5).value
    eenheid = ws.cell(row, 6).value
    ean = ws.cell(row, 7).value
    
    if is_fastener(omschrijving):
        fastener_count += 1
        fastener = {
            "id": str(artikelnummer) if artikelnummer else "",
            "naam": str(omschrijving) if omschrijving else "",
            "prijs": float(nettoprijs) if nettoprijs else 0.0,
            "eenheid": str(eenheid) if eenheid else "stuk",
            "ean": str(int(ean)) if ean and str(ean).isdigit() else str(ean) if ean else ""
        }
        
        # Optionele velden
        if brutoprijs:
            fastener["brutoprijs"] = float(brutoprijs)
        if korting:
            fastener["korting"] = float(korting)
        
        fasteners.append(fastener)

print(f"\nGevonden fasteners: {fastener_count}")
print(f"\nEerste 10 fasteners:")
print("-" * 70)
for i, f in enumerate(fasteners[:10], 1):
    print(f"{i}. {f['naam']} - €{f['prijs']} (EAN: {f['ean']})")

# Maak JSON bestand
json_filename = "fasteners.json"
with open(json_filename, 'w', encoding='utf-8') as f:
    json.dump(fasteners, f, indent=2, ensure_ascii=False)

print(f"\n" + "=" * 70)
print(f"JSON bestand opgeslagen: {json_filename}")
print(f"Totaal aantal fasteners: {len(fasteners)}")
print("=" * 70)

