#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script om Excel bestand te analyseren en JSON data te genereren voor index.html
"""

try:
    import openpyxl
except ImportError:
    print("openpyxl is niet geïnstalleerd. Installeer met: pip install openpyxl")
    exit(1)

import json
import sys

def analyse_excel(filename):
    """Analyseer Excel bestand en toon structuur"""
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        print("=" * 60)
        print(f"Excel Bestand: {filename}")
        print("=" * 60)
        print(f"\nAantal rijen: {ws.max_row}")
        print(f"Aantal kolommen: {ws.max_column}")
        
        # Lees headers (eerste rij)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col).value
            headers.append(cell_value if cell_value else f"Kolom{col}")
        
        print(f"\nKolom headers ({len(headers)} kolommen):")
        for i, header in enumerate(headers, 1):
            print(f"  {i}. {header}")
        
        # Toon eerste 5 data rijen
        print(f"\nEerste 5 data rijen:")
        print("-" * 60)
        for row in range(2, min(7, ws.max_row + 1)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row, col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Rij {row}: {' | '.join(row_data[:5])}")
            if len(row_data) > 5:
                print(f"      ... ({len(row_data) - 5} meer kolommen)")
        
        # Analyseer welke kolommen mogelijk relevant zijn
        print(f"\n" + "=" * 60)
        print("ANALYSE VOOR JSON CONVERSIE:")
        print("=" * 60)
        
        # Zoek naar mogelijke product velden
        possible_fields = {
            'naam': ['naam', 'product', 'artikel', 'omschrijving', 'item'],
            'nummer': ['nummer', 'artikelnummer', 'code', 'sku', 'id'],
            'prijs': ['prijs', 'bedrag', 'price', 'cost'],
            'categorie': ['categorie', 'groep', 'category', 'type']
        }
        
        found_fields = {}
        for field_type, keywords in possible_fields.items():
            for i, header in enumerate(headers):
                header_lower = str(header).lower() if header else ""
                for keyword in keywords:
                    if keyword in header_lower:
                        found_fields[field_type] = i
                        break
                if field_type in found_fields:
                    break
        
        print("\nGevonden velden:")
        for field_type, col_index in found_fields.items():
            print(f"  {field_type}: Kolom {col_index + 1} ({headers[col_index]})")
        
        if not found_fields:
            print("  Geen standaard velden gevonden. Alle kolommen worden gebruikt.")
        
        return ws, headers
        
    except Exception as e:
        print(f"Fout bij lezen van Excel bestand: {e}")
        return None, None

if __name__ == "__main__":
    filename = "BJC bestel artikelen Oetelaar (1).xlsx"
    analyse_excel(filename)

