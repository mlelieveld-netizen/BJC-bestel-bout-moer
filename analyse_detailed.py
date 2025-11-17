#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gedetailleerde analyse van Excel bestand
"""

import openpyxl

filename = "BJC bestel artikelen Oetelaar (1).xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb.active

print("=" * 70)
print("VOLLEDIGE ANALYSE: BJC bestel artikelen Oetelaar")
print("=" * 70)

# Headers
headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(1, col).value)

print(f"\nTOTAAL AANTAL ARTIKELEN: {ws.max_row - 1}")
print(f"AANTAL KOLOMMEN: {ws.max_column}")

print("\n" + "-" * 70)
print("KOLOM STRUCTUUR:")
print("-" * 70)
for i, header in enumerate(headers, 1):
    print(f"  {i}. {header}")

print("\n" + "-" * 70)
print("VOORBEELD DATA (eerste 5 artikelen):")
print("-" * 70)

for row in range(2, min(7, ws.max_row + 1)):
    print(f"\nArtikel {row - 1}:")
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        print(f"   {headers[col-1]}: {value}")

print("\n" + "-" * 70)
print("DATA TYPEN ANALYSE:")
print("-" * 70)

# Analyseer data types en voorbeelden
for col_idx, header in enumerate(headers, 1):
    print(f"\n{header}:")
    sample_values = []
    for row in range(2, min(10, ws.max_row + 1)):
        value = ws.cell(row, col_idx).value
        if value is not None:
            sample_values.append(str(value))
    if sample_values:
        print(f"  Voorbeelden: {', '.join(sample_values[:3])}")

print("\n" + "=" * 70)
print("MAPPING VOOR JSON CONVERSIE:")
print("=" * 70)
print("""
Voor de index.html DATA structuur hebben we nodig:
- id/nummer: Artikelnummer (kolom 1)
- naam: Artikelomschrijving (kolom 2)
- prijs: Nettoprijs (kolom 5) - dit is de prijs na korting
- categorie: (niet aanwezig, kan handmatig toegevoegd worden)
- eenheid: Verkoopprijseenheid (kolom 6)
- ean: ean-nummer (kolom 7)

Optionele velden:
- brutoprijs: Brutoprijs (kolom 3)
- korting: Kortingpercentage (kolom 4)
""")

